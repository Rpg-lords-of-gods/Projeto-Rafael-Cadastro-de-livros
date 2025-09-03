import openpyxl
import os

# Verifica se o arquivo já existe
if os.path.exists("livros.xlsx"):
    workbook = openpyxl.load_workbook("livros.xlsx")
    sheet = workbook.active
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Livros"
    sheet.append(["Título", "Autor", "Ano", "Preço", "Quantidade"])  # Cabeçalho


# ---------------- FUNÇÕES ---------------- #

def mostrar_menu():
    print("\n=== Sistema da Loja de Livros ===")
    print("1. Cadastrar livro")
    print("2. Listar livros cadastrados")
    print("3. Buscar livro por título")
    print("4. Editar livro")
    print("5. Excluir livro")
    print("6. Sair")


def cadastrar_livro():
    print("\n=== Cadastro de Livro ===")

    while True:
        titulo = input("Digite o título do livro: ").strip()
        if titulo:
            break
        print("Título inválido! Não pode ficar vazio.")

    while True:
        autor = input("Digite o autor do livro: ").strip()
        if autor:
            break
        print("Autor inválido! Não pode ficar vazio.")

    while True:
        ano = input("Digite o ano de publicação: ").strip()
        if ano.isdigit() and len(ano) == 4:
            ano = int(ano)
            break
        print("Ano inválido! Digite um número de 4 dígitos.")

    while True:
        preco = input("Digite o preço: ").strip().replace(",", ".")
        try:
            preco = float(preco)
            if preco >= 0:
                break
            else:
                print("Preço inválido! Não pode ser negativo.")
        except ValueError:
            print("Preço inválido! Digite um número válido.")

    while True:
        quantidade = input("Digite a quantidade: ").strip()
        if quantidade.isdigit():
            quantidade = int(quantidade)
            if quantidade >= 0:
                break
            else:
                print("Quantidade inválida! Não pode ser negativa.")
        else:
            print("Quantidade inválida! Digite apenas números inteiros.")

    sheet.append([titulo, autor, ano, preco, quantidade])
    workbook.save("livros.xlsx")
    print(f"\nLivro '{titulo}' cadastrado com sucesso!")


def listar_livros():
    print("\n=== Lista de Livros Cadastrados ===")

    if sheet.max_row == 1:
        print("Nenhum livro cadastrado ainda.")
        return

    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:
            print(f"{row[0]:<30} {row[1]:<20} {row[2]:<6} {row[3]:<8} {row[4]:<10}")
            print("-" * 80)
        else:
            titulo, autor, ano, preco, quantidade = row
            print(f"{titulo:<30} {autor:<20} {ano:<6} {preco:<8} {quantidade:<10}")


def buscar_livro():
    print("\n=== Buscar Livro por Título ===")
    termo = input("Digite o título ou parte do título: ").strip().lower()

    encontrados = []
    for row in sheet.iter_rows(values_only=True):
        if row[0].lower().find(termo) != -1:
            encontrados.append(row)

    if not encontrados:
        print("Nenhum livro encontrado com esse termo.")
        return

    print(f"\n{'Título':<30} {'Autor':<20} {'Ano':<6} {'Preço':<8} {'Quantidade':<10}")
    print("-" * 80)
    for livro in encontrados:
        titulo, autor, ano, preco, quantidade = livro
        print(f"{titulo:<30} {autor:<20} {ano:<6} {preco:<8} {quantidade:<10}")


def editar_livro():
    print("\n=== Editar Livro ===")
    termo = input("Digite o título do livro que deseja editar: ").strip().lower()

    encontrados = []
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        if termo in row[0].lower():
            encontrados.append((i, row))

    if not encontrados:
        print("Nenhum livro encontrado com esse termo.")
        return

    print(f"\n{'ID':<4} {'Título':<30} {'Autor':<20} {'Ano':<6} {'Preço':<8} {'Quantidade':<10}")
    print("-" * 80)
    for idx, (linha, livro) in enumerate(encontrados, start=1):
        titulo, autor, ano, preco, quantidade = livro
        print(f"{idx:<4} {titulo:<30} {autor:<20} {ano:<6} {preco:<8} {quantidade:<10}")

    while True:
        escolha = input("\nDigite o ID do livro que deseja editar: ").strip()
        if escolha.isdigit() and 1 <= int(escolha) <= len(encontrados):
            escolha = int(escolha)
            break
        print("ID inválido!")

    linha_editar = encontrados[escolha - 1][0]

    print("Digite os novos dados (deixe vazio para manter o atual):")
    for coluna, campo in enumerate(["Título", "Autor", "Ano", "Preço", "Quantidade"]):
        atual = sheet.cell(row=linha_editar, column=coluna + 1).value
        novo = input(f"{campo} (atual: {atual}): ").strip()
        if novo:
            if campo == "Ano" or campo == "Quantidade":
                if not novo.isdigit():
                    print(f"{campo} inválido! Mantendo valor atual.")
                    continue
                novo = int(novo)
            if campo == "Preço":
                try:
                    novo = float(novo.replace(",", "."))
                except:
                    print(f"{campo} inválido! Mantendo valor atual.")
                    continue
            sheet.cell(row=linha_editar, column=coluna + 1, value=novo)

    workbook.save("livros.xlsx")
    print("\nLivro atualizado com sucesso!")


def excluir_livro():
    print("\n=== Excluir Livro ===")
    termo = input("Digite o título do livro que deseja excluir: ").strip().lower()

    encontrados = []
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        if termo in row[0].lower():
            encontrados.append((i, row))

    if not encontrados:
        print("Nenhum livro encontrado com esse termo.")
        return

    print(f"\n{'ID':<4} {'Título':<30} {'Autor':<20} {'Ano':<6} {'Preço':<8} {'Quantidade':<10}")
    print("-" * 80)
    for idx, (linha, livro) in enumerate(encontrados, start=1):
        titulo, autor, ano, preco, quantidade = livro
        print(f"{idx:<4} {titulo:<30} {autor:<20} {ano:<6} {preco:<8} {quantidade:<10}")

    while True:
        escolha = input("\nDigite o ID do livro que deseja excluir: ").strip()
        if escolha.isdigit() and 1 <= int(escolha) <= len(encontrados):
            escolha = int(escolha)
            break
        print("ID inválido!")

    linha_excluir = encontrados[escolha - 1][0]

    sheet.delete_rows(linha_excluir)
    workbook.save("livros.xlsx")
    print("\nLivro excluído com sucesso!")


# ---------------- LOOP PRINCIPAL ---------------- #

while True:
    mostrar_menu()
    opcao = input("Escolha uma opção: ").strip()

    if opcao == "1":
        cadastrar_livro()
    elif opcao == "2":
        listar_livros()
    elif opcao == "3":
        buscar_livro()
    elif opcao == "4":
        editar_livro()
    elif opcao == "5":
        excluir_livro()
    elif opcao == "6":
        print("Saindo do sistema...")
        break
    else:
        print("Opção inválida! Tente novamente.")
