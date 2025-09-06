from flask import Flask, render_template, request, redirect, url_for, flash
import openpyxl
from flask import jsonify
import difflib
import os

app = Flask(__name__)
app.secret_key = "dev"  # usado para mensagens flash

EXCEL_FILE = "livros.xlsx"
HEADERS = ["Título", "Autor", "Ano", "Preço", "Quantidade"]

# Função para abrir ou criar a planilha
def get_sheet():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sh = wb.active
    else:
        wb = openpyxl.Workbook()
        sh = wb.active
        sh.title = "Livros"
        sh.append(HEADERS)
        wb.save(EXCEL_FILE)
    return wb, sh

# Função para salvar livro no Excel
def append_livro(titulo, autor, ano, preco, quantidade):
    wb, sh = get_sheet()
    sh.append([titulo, autor, ano, preco, quantidade])
    wb.save(EXCEL_FILE)

# Página inicial
@app.route("/")
def index():
    return render_template("index.html")

# Página de cadastro
@app.route("/cadastrar", methods=["GET", "POST"])
def cadastrar():
    if request.method == "POST":
        # Coletar dados do formulário
        titulo = request.form.get("titulo", "").strip()
        autor = request.form.get("autor", "").strip()
        ano = request.form.get("ano", "").strip()
        preco = request.form.get("preco", "").strip().replace(",", ".")
        quantidade = request.form.get("quantidade", "").strip()

        # Validações
        erros = []
        if not titulo: erros.append("Título é obrigatório.")
        if not autor: erros.append("Autor é obrigatório.")
        if not (ano.isdigit() and len(ano) == 4): erros.append("Ano inválido.")
        try:
            preco_val = float(preco)
            if preco_val < 0: erros.append("Preço não pode ser negativo.")
        except:
            erros.append("Preço inválido.")
        if not quantidade.isdigit() or int(quantidade) < 0:
            erros.append("Quantidade inválida.")

        # Se houver erros, exibir na tela
        if erros:
            for e in erros:
                flash(e, "erro")
            return render_template("cadastrar.html", form=request.form)

        # Se estiver tudo certo, salva no Excel
        append_livro(titulo, autor, int(ano), float(preco), int(quantidade))
        flash(f"Livro '{titulo}' cadastrado com sucesso!", "sucesso")
        return redirect(url_for("cadastrar"))

    return render_template("cadastrar.html", form={})

# Página de listagem
@app.route("/listar")
def listar():
    wb, sh = get_sheet()
    livros = []

    for i, row in enumerate(sh.iter_rows(values_only=True), start=1):
        if i == 1:  # pula cabeçalho
            continue
        livros.append(row)

    return render_template("listar.html", livros=livros, headers=HEADERS)

# Página de busca
@app.route("/buscar", methods=["GET", "POST"])
def buscar():
    resultados = []
    termo = ""
    filtro = "titulo"

    if request.method == "POST":
        termo = request.form.get("termo", "").strip().lower()
        filtro = request.form.get("filtro", "titulo")  # "titulo" ou "autor"
        wb, sh = get_sheet()

        # lista de registros
        registros = list(sh.iter_rows(values_only=True))[1:]  # pula cabeçalho

        # extrai a coluna que vamos comparar (0 = título, 1 = autor)
        col_index = 0 if filtro == "titulo" else 1
        valores = [str(r[col_index]) for r in registros]

        # pega os mais parecidos
        similares = difflib.get_close_matches(termo, valores, n=10, cutoff=0.5)

        # adiciona ao resultado os registros correspondentes
        for row in registros:
            if str(row[col_index]) in similares:
                resultados.append(row)

    return render_template("buscar.html", resultados=resultados, termo=termo, filtro=filtro, headers=HEADERS)


# Página de edição
@app.route("/editar/<int:linha>", methods=["GET", "POST"])
def editar(linha):
    wb, sh = get_sheet()
    dados = [cell.value for cell in sh[linha]]

    if request.method == "POST":
        titulo = request.form.get("titulo", "").strip()
        autor = request.form.get("autor", "").strip()
        ano = request.form.get("ano", "").strip()
        preco = request.form.get("preco", "").strip().replace(",", ".")
        quantidade = request.form.get("quantidade", "").strip()

        # Validações rápidas
        if not titulo or not autor:
            flash("Título e Autor são obrigatórios!", "erro")
            return render_template("editar.html", form=request.form, linha=linha)

        try:
            ano = int(ano)
            preco = float(preco)
            quantidade = int(quantidade)
        except:
            flash("Ano, preço ou quantidade inválidos!", "erro")
            return render_template("editar.html", form=request.form, linha=linha)

        # Atualiza a planilha
        for col, val in enumerate([titulo, autor, ano, preco, quantidade], start=1):
            sh.cell(row=linha, column=col, value=val)
        wb.save(EXCEL_FILE)

        flash("Livro atualizado com sucesso!", "sucesso")
        return redirect(url_for("listar"))

    return render_template("editar.html", form={
        "titulo": dados[0],
        "autor": dados[1],
        "ano": dados[2],
        "preco": dados[3],
        "quantidade": dados[4]
    }, linha=linha)

# Página de exclusão
@app.route("/excluir/<int:linha>")
def excluir(linha):
    wb, sh = get_sheet()
    sh.delete_rows(linha)
    wb.save(EXCEL_FILE)
    flash("Livro excluído com sucesso!", "sucesso")
    return redirect(url_for("listar"))

# Rota para sugestões (autocomplete inteligente)
@app.route("/sugestoes")
def sugestoes():
    termo = request.args.get("q", "").lower()
    filtro = request.args.get("f", "titulo")  # pode ser "titulo" ou "autor"
    wb, sh = get_sheet()

    registros = list(sh.iter_rows(values_only=True))[1:]  # pula cabeçalho
    col_index = 0 if filtro == "titulo" else 1
    valores = [str(r[col_index]) for r in registros if r[col_index]]

    if not termo:
        return jsonify([])

    # fuzzy match para sugestões
    similares = difflib.get_close_matches(termo, [v.lower() for v in valores], n=5, cutoff=0.3)

    sugestoes = []
    for v in valores:
        if v.lower() in similares:
            sugestoes.append(v)

    return jsonify(sugestoes)


if __name__ == "__main__":
    app.run(debug=True)
