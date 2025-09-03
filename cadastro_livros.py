import openpyxl
import os

# Verifica se o arquivo já existe
if os.path.exists("livros.xlsx"):
    workbook = openpyxl.load_workbook("livros.xlsx")
    sheet = workbook.active
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Livros"
    sheet.append(["Título", "Autor", "Ano", "Preço", "Quantidade"])  # Cabeçalho

# Função para cadastrar um livro com validação
def cadastrar_livro():
    print("\n=== Cadastro de Livro ===")

    # Validação do título
    while True:
        titulo = input("Digite o título do livro: ").strip()
        if titulo != "":
            break
        print("Título inválido! Não pode ficar vazio.")

    # Validação do autor
    while True:
        autor = input("Digite o autor do livro: ").strip()
        if autor != "":
            break
        print("Autor inválido! Não pode ficar vazio.")

    # Validação do ano (somente números inteiros)
    while True:
        ano = input("Digite o ano de publicação: ").strip()
        if ano.isdigit() and len(ano) == 4:
            ano = int(ano)
            break
        print("Ano inválido! Digite um número de 4 dígitos.")

    # Validação do preço (somente números, permite ponto ou vírgula)
    while True:
        preco = input("Digite o preço: ").strip().replace(",", ".")
        try:
            preco = float(preco)
            if preco >= 0:
                break
            else:
                print("Preço inválido! Não pode ser negativo.")
        except ValueError:
            print("Preço inválido! Digite um número válido.")

    # Validação da quantidade (somente números inteiros)
    while True:
        quantidade = input("Digite a quantidade: ").strip()
        if quantidade.isdigit():
            quantidade = int(quantidade)
            if quantidade >= 0:
                break
            else:
                print("Quantidade inválida! Não pode ser negativa.")
        else:
            print("Quantidade inválida! Digite apenas números inteiros.")

    # Salva no Excel
    sheet.append([titulo, autor, ano, preco, quantidade])
    workbook.save("livros.xlsx")
    print(f"\nLivro '{titulo}' cadastrado com sucesso!")

# Menu principal
def mostrar_menu():
    print("\n=== Sistema da Loja de Livros ===")
    print("1. Cadastrar livro")
    print("2. Listar livros cadastrados")
    print("3. Sair")

# Loop principal
while True:
    mostrar_menu()
    opcao = input("Escolha uma opção: ").strip()

    if opcao == "1":
        cadastrar_livro()
    elif opcao == "2":
        print("Você escolheu: Listar livros cadastrados")
        # Aqui podemos colocar a função de listagem depois
    elif opcao == "3":
        print("Saindo do sistema...")
        break
    else:
        print("Opção inválida! Tente novamente.")

def listar_livros():
    print("\n=== Lista de Livros Cadastrados ===")

    # Verifica se há apenas o cabeçalho
    if sheet.max_row == 1:
        print("Nenhum livro cadastrado ainda.")
        return

    # Mostra os livros
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:
            # Cabeçalho
            print(f"{row[0]:<30} {row[1]:<20} {row[2]:<6} {row[3]:<8} {row[4]:<10}")
            print("-" * 80)
        else:
            titulo, autor, ano, preco, quantidade = row
            print(f"{titulo:<30} {autor:<20} {ano:<6} {preco:<8} {quantidade:<10}")

